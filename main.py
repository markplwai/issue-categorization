#! /usr/bin/env python3

'''Automated issue categorization using AI.
Loads issues from an Excel file, filters them by age, takes screenshots of provided URLs,
generates issue categories ("buckets") using GPT-5 Mini, categorizes issues into these
buckets using GPT-5 Nano, and outputs the results to a CSV file.'''

# pylint: disable=W0231
# pylint: disable=W0718

### IMPORTS ###

# Standard
import asyncio
import base64
import csv
import datetime
import sys
import warnings

# Typing
from typing import Optional

# 3rd Party
from openai import OpenAI  # AI API
import openpyxl  # Excel
from playwright.async_api import async_playwright, Page  # Web automation
from pydantic import BaseModel, Field  # Structured output
from termcolor import colored  # Colored terminal output

# Suppress openpyxl warnings about data validation
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

### GLOBALS ###

CLIENT: OpenAI = OpenAI()
DEBUG = 1

### HELPER CLASSES AND FUNCTIONS ###

# Helper Classes

class Issue(BaseModel):
    '''Schema for an issue'''
    timestamp: datetime.datetime = Field(..., description="Issue creation timestamp")
    grade: int = Field(..., ge=3, le=6, description="Grade level (3-6)")
    unit: int | str = Field(..., description="Unit identifier (number 1-15 or 'Warm Up')")
    lesson: Optional[str] = Field(None, description="Lesson name")
    content_type: str = Field(..., description="Content type")
    content_item_id: Optional[str] = Field(None, description="Content item ID")
    issue_type: str = Field(..., description="Issue type ('Formatting Error', 'Duplicates', etc.)")
    severity: str = Field(..., description="Issue severity ('Minor', 'Medium', 'Major')")
    description: Optional[str] = Field(..., description="Issue description")
    screenshot_urls: Optional[str] = Field(None, description="Screenshot URLs (newline-separated)")

class BucketedIssue(Issue):
    '''Schema for an issue with categorization'''
    bucket_name: str = Field(..., description="Assigned bucket name")
    confidence: int = Field(..., ge=0, le=100, description="Confidence score (0-100)")
    rationale: str = Field(..., description="Brief rationale for the choice")

class Categorization(BaseModel):
    '''Schema for issue categorization'''
    bucket: 'Bucket' = Field(..., description="Assigned bucket name")
    confidence: int = Field(..., ge=0, le=100, description="Confidence score (0-100)")
    rationale: str = Field(..., description="Brief rationale for the choice")

class Bucket(BaseModel):
    '''Schema for an issue bucket'''
    name: str = Field(..., description="Name of the issue category ('bucket')")
    definition: str = Field(..., description="Definition of the issue bucket")
    example: str = Field(..., description="Short example of an issue in this bucket")

class BucketsResponse(BaseModel):
    '''Schema for the bucket-making response'''
    buckets: list[Bucket] = Field(..., description="List of issue buckets")

# Helper Functions

# Progress bar helper
def print_progress_bar(iteration, total, prefix='', suffix='', length=20):
    '''Prints a progress bar to the terminal.'''
    percent = int(100 * (iteration / float(total))) if total else 100
    filled_length = int(length * iteration // total) if total else length
    progress_bar = '=' * filled_length + ' ' * (length - filled_length)
    ratio = f"{iteration}/{total}"
    print(f'\r{prefix} [{progress_bar}] {percent}% ({ratio}) {suffix}', end='\r', flush=True)
    if iteration >= total:
        print()

### FUNCTIONS ###

# Issue loader
def load_issues(file_path: str) -> list[Issue]:
    """Load issues from an Excel file and return them as a list of Issue objects."""
    try:
        workbook = openpyxl.load_workbook(file_path)
        if DEBUG:
            print(
                colored("[DEBUG]", "cyan"),
                f"Loaded workbook {file_path}\n        with sheets: {workbook.sheetnames}\n",
            )

        issues: list[Issue] = []
        for sheet in workbook:
            for row in sheet.iter_rows(min_row=2, values_only=True):
                try:
                    if row[0] and row[1] and row[2] and row[4] and row[6] and row[7]:
                        issue = Issue(
                            timestamp=row[0],  # type: ignore
                            grade=row[1],  # type: ignore
                            unit=row[2],  # type: ignore
                            lesson=row[3],  # type: ignore
                            content_type=row[4],  # type: ignore
                            content_item_id=row[5],  # type: ignore
                            issue_type=row[6],  # type: ignore
                            severity=row[7],  # type: ignore
                            description=row[8],  # type: ignore
                            screenshot_urls=row[9] if len(row) > 9 else None  # type: ignore
                        )
                        issues.append(issue)

                except Exception as e:
                    print(
                        colored("[ERROR]", "red"),
                        f"Error parsing row {row} in sheet {sheet.title}: {e}\n"
                    )

        if DEBUG:
            print(colored("[DEBUG]", "cyan"), f"Loaded {len(issues)} issues.\n")
            print(colored("[DEBUG]", "cyan"), f"Sample issue: {issues[0]}\n")

        return issues

    except Exception as e:
        print(
            colored("[ERROR]", "red", attrs=["bold"]),
            f"Error loading issues from {file_path}: {e}\n"
        )
        sys.exit(1)

# Issue filter
def filter_issues(issues: list[Issue], age: int) -> list[Issue]:
    """Filter issues based on their age in days (keep those within `age` days)."""
    skipped_count = 0
    try:
        filtered_issues: list[Issue] = []
        current_date = datetime.datetime.now()

        for issue in issues:
            if not issue.timestamp:  # type: ignore
                skipped_count += 1
                continue
            try:
                created_at = issue.timestamp  # type: ignore
                if isinstance(created_at, datetime.date) and not isinstance(
                    created_at, datetime.datetime
                ):
                    created_at = datetime.datetime.combine(
                        created_at, datetime.time.min
                    )
                issue_age = (current_date - created_at).days
                if issue_age <= age:
                    filtered_issues.append(issue)
            except Exception as e:
                print(
                    colored("[ERROR]", "red"),
                    f"Error processing issue {issue}: {e}\n"
                )
                continue

        if DEBUG:
            print(
                colored("[DEBUG]", "cyan"),
                f"Filtered issues: {len(filtered_issues)} "
                f"out of {len(issues)} (skipped {skipped_count})\n",
            )

        return filtered_issues

    except Exception as e:
        print(
            colored("[ERROR]", "red", attrs=["bold"]),
            f"Error filtering issues: {e}\n"
        )
        sys.exit(1)

# Screenshot URL extractor
def extract_urls(cell: Optional[str]) -> list[str]:
    """Split a col-J cell (may have newlines) into a list of URLs."""
    if not cell:
        return []
    return [u.strip() for u in str(cell).splitlines() if u.strip()]

# Screenshot taker
async def screenshot_urls(urls: list[str], page: Page) -> list[str]:
    """
    Given a list of URLs, visit them in a shared browser and return
    base64-encoded full-page screenshots.
    """
    urls = [u.strip() for u in urls if u.strip()]
    results: list[Optional[str]] = [None] * len(urls)

    for idx, url in enumerate(urls):
        try:
            await page.goto(url, wait_until="load", timeout=20000)
            shot = await page.screenshot(full_page=True, type="jpeg")
            results[idx] = base64.b64encode(shot).decode()
        except Exception as e:
            if "net::ERR_NAME_NOT_RESOLVED" in str(e):
                print(colored("[WARNING]", "yellow"), f"Invalid URL {url}, skipping")
            elif "TimeoutError" in str(e):
                print(colored("[WARNING]", "yellow"), f"Timeout loading {url}, skipping")
            else:
                print(colored("[ERROR]", "red"), f"Screenshot failed for {url}: {e}")

    return [r for r in results if r]

# Convert issues to API input format
async def create_inputs(issues: list[Issue], concurrency: int = 20) -> list[dict]:
    """
    Convert issues into Responses API input format.
    Each issue -> one user message with text + all screenshots.
    """
    inputs = []
    total = len(issues)

    semaphore = asyncio.Semaphore(concurrency)
    lock = asyncio.Lock()
    completed = {'count': 0}  # Use a dict for mutability in closure

    async def summarize_image(text: str, b64: str) -> str:
        """
        Use GPT-5 Nano to summarize the image (with the relevant issue text).
        Returns the summary string.
        """
        try:
            system = (
                "You are an expert at summarizing screenshots for content issues.\n"
                "Given the following issue description and a screenshot, "
                "provide a concise summary of what the screenshot shows, "
                "focusing on details relevant to the issue.\n"
                "Keep it as short as possible! Don't summarize the issue, just the image."
            )
            user_content = [
                {"type": "input_text", "text": text},
                {"type": "input_image", "image_url": f"data:image/jpeg;base64,{b64}"}
            ]
            response = CLIENT.responses.create(
                model="gpt-5-nano",
                input=[
                    {"role": "system", "content": system},
                    {"role": "user", "content": user_content}  # type: ignore
                ]
            )
            return response.output_text
        except Exception as e:
            print(colored("[ERROR]", "red", attrs=["bold"]), f"Image summarization failed: {e}\n")
            return "[Image summary unavailable]"

    async def build_message_with_semaphore(issue):
        async with semaphore:
            page = await browser.new_page()

            # Text block
            text = (
                f"Lesson: {issue.lesson} | Content Type: {issue.content_type} | "
                f"Issue Severity: {issue.severity} | Issue Description: {issue.description}"
            )

            # Screenshots from col J
            urls = extract_urls(issue.screenshot_urls)
            b64s = []
            imgs = []
            if urls:
                b64s = await screenshot_urls(urls, page)
                # Summarize each image with GPT-5 Nano
                for b64 in b64s:
                    summary = await summarize_image(text, b64)
                    imgs.append({
                        "type": "input_text",
                        "text": "Image summary: " + summary
                    })

            # Progress bar update
            async with lock:
                completed['count'] += 1
                print_progress_bar(completed['count'], total, prefix="          ")

            await page.close()

            return {"role": "user", "content": [{"type": "input_text", "text": text}] + imgs}

    # Run all message builds concurrently, but limit concurrency with semaphore
    try:
        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=True)
            inputs = await asyncio.gather(
                *(build_message_with_semaphore(issue) for issue in issues)
            )
            await browser.close()

        return inputs

    except Exception as e:
        print(
            colored("[ERROR]", "red", attrs=["bold"]),
            f"Error creating inputs and uploading to vector store: {e}\n"
        )
        sys.exit(1)

# AI bucket-maker
def ai_make_buckets(inputs: list[dict]) -> list[Bucket]:
    """Generate new, targeted buckets for issues using OpenAI's GPT-5 Mini."""
    system = (
        """You are an expert at creating issue categories ('buckets') for content issues.
        You are given a list of issues with descriptions and image summaries.
        Your task is to create a concise list of distinct issue categories ('buckets') that
        effectively cover the range of issues provided.
        Each bucket should have:
        1. A clear, descriptive name (3-5 words).
        2. A brief definition (1-2 sentences).
        3. A short example of an issue that would fit in this bucket.
        Guidelines:
        - Create as many buckets as needed to cover the issues, but avoid redundancy.
        - Ensure buckets are distinct and non-overlapping.
        - Buckets should be as specific as possible to the issues provided, but not too narrow.
        - Buckets shouldn't be too broad either.
        - If an issue does not fit well into any bucket, create a new bucket for it.
        """
    )
    try:
        response = CLIENT.responses.parse(
            model="gpt-5-mini",
            input=[
                {"role": "system", "content": system}  # type: ignore
            ] + inputs,
            reasoning={"effort": "high"},
            text_format=BucketsResponse,
        )

        return response.output_parsed.buckets  # type: ignore

    except Exception as e:
        print(
            colored("[ERROR]", "red", attrs=["bold"]),
            f"Error generating buckets: {e}\n"
        )
        sys.exit(1)


def ai_categorize_issues(issue_inputs: list[dict], buckets: list[Bucket]) -> list[Categorization]:
    """Categorize issues into the provided buckets using OpenAI's GPT-5 Nano.
    Returns a category, confidence score, and rationale for each issue."""
    try:
        system = (
            "You are an expert categorizer of content issues.\n"
            "You are given a list of issue categories ('buckets') with definitions.\n"
            "You are to assign the issue given to you to one of these buckets.\n"
            "For each issue, provide:\n"
            "1. The most appropriate bucket name.\n"
            "2. A confidence score (0-100) for your assignment.\n"
            "3. A brief rationale for your choice.\n"
            "Guidelines:\n"
            "- Choose the bucket that best fits the issue.\n"
            "- If none fit well, choose the closest match.\n"
        )

        categorizations: list[Categorization] = []
        total = len(issue_inputs)
        for idx, issue in enumerate(issue_inputs):
            response = CLIENT.responses.parse(
                model="gpt-5-nano",
                input=[
                    {"role": "system", "content": system},
                    {"role": "system", "content": f"Available buckets: \n\n{buckets}"},
                    {"role": "user", "content": issue["content"]}
                ],
                text_format=Categorization,
            )

            categorizations.append(response.output_parsed)  # type: ignore
            print_progress_bar(idx + 1, total, prefix="          ")

        return categorizations

    except Exception as e:
        print(
            colored("[ERROR]", "red", attrs=["bold"]),
            f"Error categorizing issues: {e}\n"
        )
        sys.exit(1)

# Main function
def main():
    """Main function to load, filter, and process issues."""

    ### LOAD ISSUES ###

    print(colored("[PROGRESS]", "light_green", attrs=["bold"]), "Loading issues...", end="\r")
    issues: list[Issue] = load_issues("issues.xlsx")
    print(colored("[PROGRESS]", "light_green", attrs=["bold"]),
          "Loading issues...", colored("Done.", "light_green"),
          f"Total issues loaded: {len(issues)}.")

    ### FILTER ISSUES ###

    print(colored("[PROGRESS]", "light_green", attrs=["bold"]), "Filtering issues...", end="\r")
    filtered_issues: list[Issue] = filter_issues(issues, age=30)
    if not filtered_issues:
        print(colored("[WARNING]", "yellow", attrs=["bold"]), "All issues filtered out! Exiting.")
        return

    print(colored("[PROGRESS]", "light_green", attrs=["bold"]),
          "Filtering issues...", colored("Done.", "light_green"),
          f"{len(filtered_issues)} of {len(issues)} issues kept.")

    if DEBUG:
        print(colored("[DEBUG]", "cyan"), f"First filtered issue: {filtered_issues[0]}\n")
        print(colored("[DEBUG]", "cyan"), f"Total filtered issues: {len(filtered_issues)}\n")

    ### PROCESS ISSUES ###

    print(colored("[PROGRESS]", "light_green", attrs=["bold"]),
          "Building API inputs (with screenshots)...")
    if DEBUG:
        inputs: list[dict] = asyncio.run(create_inputs(filtered_issues[:5]))  # First 5
        print(colored("[DEBUG]", "cyan"), f"First input: {str(inputs[0])[:1000]}...\n")
        print(colored("[DEBUG]", "cyan"), f"Total inputs: {len(inputs)}\n")
    else:
        inputs: list[dict] = asyncio.run(create_inputs(filtered_issues))

    print(colored("[PROGRESS]", "light_green", attrs=["bold"]),
          "Building API inputs (with screenshots)...", colored("Done.", "light_green"))

    ### AI ISSUE CATEGORY CREATION ###

    print(colored("[PROGRESS]", "light_green", attrs=["bold"]),
          "Sending issues to bucket-defining assistant...", end="\r")
    buckets: list[Bucket] = ai_make_buckets(inputs)
    print(colored("[PROGRESS]", "light_green", attrs=["bold"]),
          "Sending issues to bucket-defining assistant...", colored("Done.", "light_green"), 
          f"{len(buckets)} buckets defined.")
    if DEBUG:
        print(colored("[DEBUG]", "cyan"), f"Buckets: {buckets}\n")

    ### AI ISSUE CATEGORIZATION W/ NEW CATEGORIES ###

    print(colored("[PROGRESS]", "light_green", attrs=["bold"]),
          "Categorizing issues into buckets...", end="\r")
    categorizations: list[Categorization] = ai_categorize_issues(inputs, buckets)
    bucketed_issues: list[BucketedIssue] = [
        BucketedIssue(
            **issue.model_dump(),
            bucket_name=cat.bucket.name,
            confidence=cat.confidence,
            rationale=cat.rationale
        ) for issue, cat in zip(filtered_issues, categorizations)
    ]
    print(colored("[PROGRESS]", "light_green", attrs=["bold"]),
          "Categorizing issues into buckets...", colored("Done.", "light_green"))
    if DEBUG:
        print(colored("[DEBUG]", "cyan"), f"First categorization: {categorizations[0]}\n")
        print(colored("[DEBUG]", "cyan"), f"Total categorizations: {len(categorizations)}\n")

    ### OUTPUT CSV CREATION ###

    print(colored("[PROGRESS]", "light_green", attrs=["bold"]),
          "Writing results to output.csv...", end="\r")
    with open('output.csv', mode='w', newline='', encoding='utf-8') as file:
        writer = csv.writer(file)
        writer.writerow(bucketed_issues[0].model_dump().keys())  # Header
        for issue in bucketed_issues:
            writer.writerow(issue.model_dump().values())

        file.close()

    print(colored("[PROGRESS]", "light_green", attrs=["bold"]),
          "Writing results to output.csv...", colored("Done.", "light_green"))

# Entry point
if __name__ == "__main__":
    main()
